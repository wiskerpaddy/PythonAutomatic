#####################################################################
#  ・エクセルのセルの内容をもとに、リネームするプログラム　            #
#####################################################################

from pathlib import Path
from openpyxl import load_workbook

try:
    targetPath = Path("work")
    # リネーム対象のフォルダが存在しなければ例外をスロー
    if not (Path(targetPath).exists()):
        print("『" + targetPath.name + "』が存在しません。")
        raise FileExistsError()  
    
    # ①フォルダー内のファイル一覧を取得
    path_list = targetPath.iterdir()

    for path in path_list:
        # ②-1 ファイル名取得
        pathName = path.name
        # ②-2 ブックを読み込み
        wb = load_workbook(f"work/{pathName}")
        # ②-3 シートを読み込み
        ws = wb.worksheets[0]
        # ③ セルの値を読み込み
        no = ws["C3"].value
        name = ws["C4"].value
        # ④-1 読み込んだ値を使用してファイル名をリネーム（リネーム後のパスオブジェクトを作成）
        pathAfter = Path(f"./work/{no}{name}.xlsx")
        # ④-2 読み込んだ値を使用してファイル名をリネーム（リネーム実行）
        path.rename(pathAfter)
    
except FileExistsError as e:
    print("処理を終了します。")

except Exception as e:
    print(e)
    print("異常が発生しました。処理を終了します。")